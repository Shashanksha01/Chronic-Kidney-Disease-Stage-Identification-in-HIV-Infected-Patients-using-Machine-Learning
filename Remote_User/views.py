from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl
import pandas as pd
import numpy as np
import re
from sklearn.ensemble import VotingClassifier
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score

from sklearn.tree import DecisionTreeClassifier

# Create your views here.
from Remote_User.models import ClientRegister_Model,kidney_model,kidney_disease_model,detection_ratio_model,detection_accuracy_model


def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            kidney_model.objects.all().delete()
            kidney_disease_model.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        kidney_model.objects.create(
        id1= active_sheet.cell(r, 1).value,
        age= active_sheet.cell(r, 2).value,
        bp= active_sheet.cell(r, 3).value,
        sg= active_sheet.cell(r, 4).value,
        al= active_sheet.cell(r, 5).value,
        su= active_sheet.cell(r, 6).value,
        rbc= active_sheet.cell(r, 7).value,
        pc= active_sheet.cell(r, 8).value,
        pcc= active_sheet.cell(r, 9).value,
        ba= active_sheet.cell(r, 10).value,
        bgr= active_sheet.cell(r, 11).value,
        bu= active_sheet.cell(r, 12).value,
        sc= active_sheet.cell(r, 13).value,
        sod= active_sheet.cell(r, 14).value,
        pot= active_sheet.cell(r, 15).value,
        hemo= active_sheet.cell(r, 16).value,
        pcv= active_sheet.cell(r, 17).value,
        wc= active_sheet.cell(r, 18).value,
        rc= active_sheet.cell(r, 19).value,
        htn= active_sheet.cell(r, 20).value,
        dm= active_sheet.cell(r, 21).value,
        cad= active_sheet.cell(r, 22).value,
        appet= active_sheet.cell(r, 23).value,
        pe= active_sheet.cell(r, 24).value,
        ane= active_sheet.cell(r, 25).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def  Prediction_Of_Kdney_Disease(request):
    if request.method == "POST":

        if request.method == "POST":
            Readingid = request.POST.get('Readingid')
            age= request.POST.get('age')
            bp= request.POST.get('bp')
            sg= request.POST.get('sg')
            al= request.POST.get('al')
            su= request.POST.get('su')
            rbc= request.POST.get('rbc')
            pc= request.POST.get('pc')
            pcc= request.POST.get('pcc')
            ba= request.POST.get('ba')
            bgr= request.POST.get('bgr')
            bu= request.POST.get('bu')
            sc= request.POST.get('sc')
            sod= request.POST.get('sod')
            pot= request.POST.get('pot')
            hemo= request.POST.get('hemo')
            pcv= request.POST.get('pcv')
            wc= request.POST.get('wc')
            rc= request.POST.get('rc')
            htn= request.POST.get('htn')
            dm= request.POST.get('dm')
            cad= request.POST.get('cad')
            appet= request.POST.get('appet')
            pe= request.POST.get('pe')
            ane= request.POST.get('ane')

            df = pd.read_csv('kidney_disease.csv', encoding='latin-1')
            df
            df.columns

            def apply_measure(pottacium):
                if pottacium >= 3.6 and pottacium <= 5.2:
                    return 0  # status = "Negative"
                elif pottacium >= 5.2 or pottacium <= 3.6:
                    return 1  # status = "Positive"

            df['results'] = df['pot'].apply(apply_measure)

            cv = CountVectorizer(lowercase=False, strip_accents='unicode', ngram_range=(1, 1))
            X = df['id']
            y = df['results']

            X = cv.fit_transform(df['id'].apply(lambda X: np.str_(X)))

            models = []
            from sklearn.model_selection import train_test_split
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.20)
            X_train.shape, X_test.shape, y_train.shape

            print("Naive Bayes")

            from sklearn.naive_bayes import MultinomialNB
            NB = MultinomialNB()
            NB.fit(X_train, y_train)
            predict_nb = NB.predict(X_test)
            naivebayes = accuracy_score(y_test, predict_nb) * 100
            print(naivebayes)
            print(confusion_matrix(y_test, predict_nb))
            print(classification_report(y_test, predict_nb))
            models.append(('naive_bayes', NB))

            # SVM Model
            print("SVM")
            from sklearn import svm
            lin_clf = svm.LinearSVC()
            lin_clf.fit(X_train, y_train)
            predict_svm = lin_clf.predict(X_test)
            svm_acc = accuracy_score(y_test, predict_svm) * 100
            print(svm_acc)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, predict_svm))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, predict_svm))
            models.append(('svm', lin_clf))

            print("Logistic Regression")

            from sklearn.linear_model import LogisticRegression
            reg = LogisticRegression(random_state=0, solver='lbfgs').fit(X_train, y_train)
            y_pred = reg.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, y_pred) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, y_pred))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, y_pred))
            models.append(('logistic', reg))

            print("Decision Tree Classifier")
            dtc = DecisionTreeClassifier()
            dtc.fit(X_train, y_train)
            dtcpredict = dtc.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, dtcpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, dtcpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, dtcpredict))
            models.append(('DecisionTreeClassifier', dtc))

            print("SGD Classifier")
            from sklearn.linear_model import SGDClassifier
            sgd_clf = SGDClassifier(loss='hinge', penalty='l2', random_state=0)
            sgd_clf.fit(X_train, y_train)
            sgdpredict = sgd_clf.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, sgdpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, sgdpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, sgdpredict))
            models.append(('SGDClassifier', sgd_clf))

            print("KNeighborsClassifier")
            from sklearn.neighbors import KNeighborsClassifier
            kn = KNeighborsClassifier()
            kn.fit(X_train, y_train)
            knpredict = kn.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, knpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, knpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, knpredict))
            models.append(('KNeighborsClassifier', dtc))

            classifier = VotingClassifier(models)
            classifier.fit(X_train, y_train)
            y_pred = classifier.predict(X_test)

            review_data = [Readingid]
            vector1 = cv.transform(review_data).toarray()
            predict_text = classifier.predict(vector1)

            pred = str(predict_text).replace("[", "")
            pred1 = pred.replace("]", "")

            prediction = int(pred1)

            if (prediction == 0):
                val = 'Negative Stage'
            elif (prediction == 1):
                val = 'Positive Stage'


        return render(request, 'RUser/Prediction_Of_Kdney_Disease.html', {'objs': val})
    return render(request, 'RUser/Prediction_Of_Kdney_Disease.html')




